import openpyxl as xl

print('Hello, Welcome to my project')
    #accessing the worksheets
p = xl.load_workbook("products.xlsx")
d = xl.load_workbook("discount.xlsx")

p_sheet = p['Sheet1']
d_sheet = d['Sheet1']

p.save("temp.xlsx")
t = xl.load_workbook('temp.xlsx')
t_sheet = t['Sheet1']
    #loop to acess discount, price, quantity
for i in range(2, p_sheet.max_row + 1):
    pc = p_sheet.cell(i, 1)
    for j in range(2, d_sheet.max_row + 1):
        dc = d_sheet.cell(j, 1)
        if pc.value == dc.value:
            tc = t_sheet.cell(i, 3)
            pr = p_sheet.cell(i, 3)
            di = d_sheet.cell(j, 2)
            qc = p_sheet.cell(i, 4)
            tc.value = pr.value * (di.value*0.01) * qc.value #calculation of total price
    #erasing unwanted data in temp file which is copy of products
for i in range(1, t_sheet.max_row + 1):
    m = t_sheet.cell(i, 4)
    m.value = ''
h = 0 #finding the highest of total price
for i in range(2, t_sheet.max_row + 1):
    k = t_sheet.cell(i, 3)
    h = k.value
    if h < k.value:
        h = k.value
print('highest total price is :', h)
t.save('total_price.xlsx')
